from librouteros import connect
from openpyxl import Workbook
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import time
import os
from dotenv import load_dotenv


load_dotenv()

MIKROTIK_HOST = os.getenv("MIKROTIK_HOST")
MIKROTIK_USERNAME = os.getenv("MIKROTIK_USERNAME")
MIKROTIK_PASSWORD = os.getenv("MIKROTIK_PASSWORD")
MIKROTIK_PORT = os.getenv("MIKROTIK_PORT")

EMAIL_SENDER = os.getenv("EMAIL_SENDER")
EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD")
# EMAIL_RECEIVERS= os.getenv("EMAIL_RECEIVERS")
EMAIL_RECEIVERS= ["xrifnor@gmail.com", "glenncervantesgc4@gmail.com"]

api = connect(
    host=MIKROTIK_HOST,
    username=MIKROTIK_USERNAME,
    password=MIKROTIK_PASSWORD,
    port=MIKROTIK_PORT
)



def get_ppp_secrets():
    secrets = api("/ppp/secret/print")
    secret_list = []
    id = 1
    
    for s in secrets:
        secret_list.append({
            "id": id,
            "name": s.get("name"),
            "password": s.get("password"),
            "profile": s.get("profile"),
            "service": s.get("service"),
            "comment": s.get("comment")
        })
        id = id + 1
    return secret_list

# get_ppp_secrets()


def get_ppp_active():
    active = api("/ppp/active/print")
    active_list = []
    id = 1
    for a in active:
        active_list.append({
            "id": id,
            "name": a.get("name"),
            "address": a.get("address"),
            "uptime": a.get("uptime")
        })
        id = id + 1
    return active_list

# get_ppp_active()

def determine_disconnected(secret_list, active_list):
        active_names = {a["name"] for a in active_list}
        
        disconnected = []
        for s in secret_list:
            if s["name"] not in active_names:
                disconnected.append(s)
                
        return disconnected
    
def save_to_excel(secret_list, active_list, disconnected_list):
    wb = Workbook()
    
    # Sheet 1 for all clients
    ws1 = wb.active
    ws1.title = "Client List"
    ws1.append(["No","Name", "Password", "Profile", "Service", "Comment"])
    
    for item in secret_list:
        ws1.append([
            item["id"],
            item["name"],
            item["password"],
            item["profile"],
            item["service"],
            item["comment"]
        ])    

    # Sheet 2 for active
    ws2 = wb.create_sheet("Active")
    ws2.append(["No", "Name", "Address", "Uptime"])
    for item in active_list:
        ws2.append([
            item["id"],
            item["name"],
            item["address"],
            item["uptime"]
        ])
        
    # Sheet 3 for Disconnected
    ws3 = wb.create_sheet("Disconnected")
    ws3.append(["No", "Name", "Comment"])
    id = 1
    for item in disconnected_list:
        ws3.append([
            id,
            item["name"],
            item["comment"]
        ])
        id = id + 1
    
    wb.save("CITS_Client_Report.xlsx")
    print("Saved: CITS_Client_Report.xlsx")
    
def send_email(subject, body, to_emails, attachment_path=None): 
    msg = MIMEMultipart()
    msg["From"] = EMAIL_SENDER
    msg["To"] =  ', '.join(to_emails)
    msg["Subject"] = subject
    
    msg.attach(MIMEText(body, 'plain'))
    
    if attachment_path and os.path.exists(attachment_path):
        with open(attachment_path, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f'attachment; filename="{attachment_path}"')
        msg.attach(part)
    
    # Connect to 
    with smtplib.SMTP("smtp.gmail.com", 587) as server:
        # server.ehlo()
        server.starttls()
        # server.ehlo()
        server.login(EMAIL_SENDER, EMAIL_PASSWORD)
        server.send_message(msg)
        
    print("Email notification sent!")
    
# Main automation task
def run_mikrotik_task():
    print("Fetching data....")
    secrets = get_ppp_secrets()
    active = get_ppp_active()
    disconnected = determine_disconnected(secrets, active)
    # print(disconnected_list) 
    save_to_excel(secrets, active, disconnected)
    
    if disconnected:
        disconnected_names = [d["name"] for d in disconnected]
        subject = "CITS NETWORK ALERT: Disconnected Users"
        body = f"The following {len(disconnected_names)} users are disconnected:\n\n" + "\n".join(disconnected_names)
        send_email(subject, body, EMAIL_RECEIVERS, attachment_path="CITS_Client_Report.xlsx")
    else:
        print("No Disconnected users")


# Auto run every 1-minute
seconds = 60 * 60 * 8
while True:
    try:
        run_mikrotik_task()
    except Exception as e:
        print("Error:", e)
            
    print("Waiting 60 minutes for the next run....\n")
    print(seconds)
    time.sleep(seconds)